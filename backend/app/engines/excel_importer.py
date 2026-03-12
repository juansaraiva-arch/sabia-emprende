"""
Motor de Importacion Excel — Mi Director Financiero PTY
Importa transacciones desde archivos .xlsx al Libro Diario.
Valida columnas, formatos y partida doble antes de insertar.
"""
import io
import logging
from datetime import date, datetime

import openpyxl

from app.database import get_supabase
from app.engines.accounting_engine import validate_journal_entry

logger = logging.getLogger(__name__)

# Columnas requeridas en el Excel (case-insensitive, se normalizan)
COLUMNAS_REQUERIDAS = ["fecha", "descripcion", "cuenta_codigo", "debe", "haber"]

# Mapeo de variantes de nombre de columna aceptadas
COLUMN_ALIASES = {
    "date": "fecha",
    "entry_date": "fecha",
    "description": "descripcion",
    "desc": "descripcion",
    "account_code": "cuenta_codigo",
    "account": "cuenta_codigo",
    "codigo": "cuenta_codigo",
    "codigo_cuenta": "cuenta_codigo",
    "debit": "debe",
    "credit": "haber",
    "credito": "haber",
    "debito": "debe",
    "referencia": "referencia",
    "reference": "referencia",
    "ref": "referencia",
}


def _normalize_header(h: str) -> str:
    """Normaliza un nombre de columna."""
    h = h.strip().lower().replace(" ", "_")
    return COLUMN_ALIASES.get(h, h)


def _parse_date(val) -> str | None:
    """Convierte un valor de celda Excel a string YYYY-MM-DD."""
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def importar_excel(archivo_bytes: bytes, society_id: str, user_id: str) -> dict:
    """
    Importa transacciones desde un archivo .xlsx.

    Agrupa filas por (fecha, descripcion, referencia) para formar asientos.
    Cada grupo debe cumplir partida doble (DEBE == HABER).

    Returns: {"importados": int, "errores": list, "asientos_creados": int}
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes), read_only=True)
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo Excel: {str(e)[:200]}")

    ws = wb.active
    if ws is None:
        raise ValueError("El archivo Excel no tiene hojas de calculo.")

    # Leer y normalizar headers
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [_normalize_header(str(c or "")) for c in first_row]

    # Validar columnas requeridas
    faltantes = [c for c in COLUMNAS_REQUERIDAS if c not in headers]
    if faltantes:
        raise ValueError(
            f"Columnas faltantes en el Excel: {faltantes}. "
            f"Se requieren: {COLUMNAS_REQUERIDAS}. "
            f"Columnas encontradas: {headers}"
        )

    # Leer filas y validar datos
    filas_validas = []
    errores = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        fila = dict(zip(headers, row))

        # Validar fecha
        fecha_str = _parse_date(fila.get("fecha"))
        if not fecha_str:
            errores.append({"fila": row_num, "error": f"Fecha invalida: {fila.get('fecha')}"})
            continue

        # Validar cuenta
        cuenta = str(fila.get("cuenta_codigo", "")).strip()
        if not cuenta:
            errores.append({"fila": row_num, "error": "Falta codigo de cuenta"})
            continue

        # Parsear montos
        try:
            debe = float(fila.get("debe") or 0)
            haber = float(fila.get("haber") or 0)
        except (ValueError, TypeError):
            errores.append({"fila": row_num, "error": "Montos debe/haber no son numeros validos"})
            continue

        if debe < 0 or haber < 0:
            errores.append({"fila": row_num, "error": "Montos negativos no permitidos"})
            continue

        if debe == 0 and haber == 0:
            errores.append({"fila": row_num, "error": "Debe o Haber debe ser mayor a 0"})
            continue

        descripcion = str(fila.get("descripcion", "")).strip() or "Importado desde Excel"
        referencia = str(fila.get("referencia", "")).strip() if "referencia" in headers else ""

        filas_validas.append({
            "fila_excel": row_num,
            "fecha": fecha_str,
            "descripcion": descripcion,
            "cuenta_codigo": cuenta,
            "debe": debe,
            "haber": haber,
            "referencia": referencia,
        })

    wb.close()

    if not filas_validas:
        raise ValueError(
            f"No hay filas validas para importar. "
            f"Se encontraron {len(errores)} errores."
        )

    # Agrupar filas por (fecha, descripcion, referencia) para formar asientos
    grupos: dict[tuple, list] = {}
    for fila in filas_validas:
        key = (fila["fecha"], fila["descripcion"], fila["referencia"])
        grupos.setdefault(key, []).append(fila)

    # Validar cuentas contra el plan de cuentas
    db = get_supabase()
    cuentas_res = (
        db.table("chart_of_accounts")
        .select("id, account_code")
        .eq("society_id", society_id)
        .execute()
    )
    code_to_id = {a["account_code"]: a["id"] for a in (cuentas_res.data or [])}

    asientos_creados = 0

    for (fecha, descripcion, referencia), lineas in grupos.items():
        # Verificar que todas las cuentas existen
        cuentas_invalidas = [
            l["cuenta_codigo"] for l in lineas
            if l["cuenta_codigo"] not in code_to_id
        ]
        if cuentas_invalidas:
            for l in lineas:
                if l["cuenta_codigo"] in cuentas_invalidas:
                    errores.append({
                        "fila": l["fila_excel"],
                        "error": f"Cuenta {l['cuenta_codigo']} no existe en el plan de cuentas",
                    })
            continue

        # Validar partida doble
        lines_data = [
            {"account_code": l["cuenta_codigo"], "debe": l["debe"], "haber": l["haber"]}
            for l in lineas
        ]
        validation = validate_journal_entry(lines_data)
        if not validation["valid"]:
            for l in lineas:
                errores.append({
                    "fila": l["fila_excel"],
                    "error": f"Asiento no cuadra: {'; '.join(validation['errors'])}",
                })
            continue

        # Insertar asiento via RPC atomico
        parsed_date = datetime.strptime(fecha, "%Y-%m-%d")
        rpc_lines = [
            {
                "account_code": l["cuenta_codigo"],
                "description": l.get("descripcion", ""),
                "debe": l["debe"],
                "haber": l["haber"],
            }
            for l in lineas
        ]

        try:
            res = db.rpc("crear_asiento_contable", {
                "p_society_id": society_id,
                "p_entry_date": fecha,
                "p_description": descripcion,
                "p_reference": referencia or None,
                "p_source": "csv",
                "p_lines": rpc_lines,
                "p_created_by": user_id,
            }).execute()

            if res.data:
                asientos_creados += 1
            else:
                for l in lineas:
                    errores.append({"fila": l["fila_excel"], "error": "Error al insertar en BD"})
        except Exception as e:
            error_msg = str(e)[:200]
            logger.error("excel_import_error", extra={"society_id": society_id, "error": error_msg})
            for l in lineas:
                errores.append({"fila": l["fila_excel"], "error": f"Error BD: {error_msg}"})

    logger.info(
        "excel_import_complete",
        extra={
            "society_id": society_id,
            "asientos_creados": asientos_creados,
            "filas_validas": len(filas_validas),
            "errores": len(errores),
        },
    )

    return {
        "importados": len(filas_validas) - sum(1 for e in errores),
        "asientos_creados": asientos_creados,
        "errores": errores,
        "total_filas": len(filas_validas) + len(errores),
    }
